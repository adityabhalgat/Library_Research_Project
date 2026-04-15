"""Prompt table loader for parameter-analysis prompt matrices."""

from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List


SUPPORTED_TECHNIQUES = ["monolithic", "cot", "expert", "cove", "rcot", "two_model"]

_TECHNIQUE_ALIASES = {
    "monolithic": "monolithic",
    "mono": "monolithic",
    "cot": "cot",
    "chainofthought": "cot",
    "chain_of_thought": "cot",
    "expert": "expert",
    "expertarchitect": "expert",
    "cove": "cove",
    "chainofverification": "cove",
    "chain_of_verification": "cove",
    "rcot": "rcot",
    "reversechainofthought": "rcot",
    "reverse_chain_of_thought": "rcot",
    "twomodel": "two_model",
    "two_model": "two_model",
}

_CRITERION_HEADER_ALIASES = {
    "criterion",
    "criteria",
    "parameter",
    "parameters",
    "criteria/parameter",
    "criterion/parameter",
    "criteria_parameter",
    "criterion_parameter",
    "question",
    "c",
}

_TECHNIQUE_HEADER_ALIASES = {
    "prompt_technique",
    "prompttechnique",
    "prompting_technique",
    "promptingtechnique",
    "technique",
    "strategy",
    "prompt_type",
    "prompttype",
}

_PROMPT_HEADER_ALIASES = {
    "prompt",
    "prompts",
    "instruction",
    "instructions",
    "template",
    "prompt_text",
    "prompttext",
}


def _normalize_text(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def _normalize_header(value: object) -> str:
    lowered = _normalize_text(value).lower()
    return re.sub(r"[^a-z0-9_]+", "", lowered.replace("-", "_").replace(" ", "_"))


def _canonical_criterion(value: str) -> str:
    return " ".join(str(value).strip().lower().split())


def _normalize_technique(value: str) -> str:
    return _TECHNIQUE_ALIASES.get(_normalize_header(value), "")


def _split_prompt_steps(cell_value: object) -> List[str]:
    """Split one cell into one or many prompt steps."""
    text = str(cell_value or "").strip()
    if not text:
        return []

    # Support JSON list cells for explicit multi-step prompts.
    if text.startswith("[") and text.endswith("]"):
        try:
            parsed = json.loads(text)
            if isinstance(parsed, list):
                return [str(item).strip() for item in parsed if str(item).strip()]
        except Exception:
            pass

    # Preferred separators for stable parsing.
    separators = [
        r"\n\s*---+\s*\n",
        r"\n\s*===+\s*\n",
        r"\n\s*<<<\s*STEP\s*>>>\s*\n",
        r"\s*\|\|\s*",
    ]
    for separator in separators:
        parts = [p.strip() for p in re.split(separator, text) if p.strip()]
        if len(parts) > 1:
            return parts

    # Secondary parser: "Prompt 1:", "Step 2:" labels in one cell.
    marker_regex = re.compile(r"(?im)^\s*(?:prompt|step)\s*\d+\s*[:\-]\s*")
    markers = list(marker_regex.finditer(text))
    if len(markers) > 1:
        parts: List[str] = []
        for index, marker in enumerate(markers):
            start = marker.end()
            end = markers[index + 1].start() if index + 1 < len(markers) else len(text)
            step = text[start:end].strip()
            if step:
                parts.append(step)
        if parts:
            return parts

    return [text]


@dataclass
class PromptMatrix:
    """Pair-wise prompts keyed by criterion and technique."""

    criteria: List[str]
    pair_prompts: Dict[str, Dict[str, List[str]]]

    def get_prompts(self, criterion: str, technique: str) -> List[str]:
        key = _canonical_criterion(criterion)
        return list(self.pair_prompts.get(key, {}).get(technique, []))

    @property
    def available_techniques(self) -> List[str]:
        techniques: List[str] = []
        seen = set()
        for criterion_prompts in self.pair_prompts.values():
            for technique in criterion_prompts:
                if technique not in seen:
                    seen.add(technique)
                    techniques.append(technique)
        return techniques


class PromptTableHandler:
    """Load prompt table from Excel/CSV into pair prompt matrix."""

    def __init__(self, table_path: str, sheet_name: str | None = None):
        self.table_path = Path(table_path)
        self.sheet_name = (sheet_name or "").strip() or None

        if not self.table_path.exists():
            raise FileNotFoundError(f"Prompt table not found: {self.table_path}")

    def load(self) -> PromptMatrix:
        suffix = self.table_path.suffix.lower()
        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            rows = self._read_excel_rows()
        elif suffix == ".csv":
            rows = self._read_csv_rows()
        else:
            raise ValueError(
                f"Unsupported prompt table format: {self.table_path.suffix}. "
                "Use .xlsx/.xlsm/.xltx/.xltm or .csv"
            )

        return self._build_matrix(rows)

    def _read_excel_rows(self) -> List[Dict[str, str]]:
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise RuntimeError(
                "openpyxl is required to read Excel prompt tables. Install it with `pip install openpyxl`."
            ) from exc

        workbook = load_workbook(self.table_path, data_only=True)
        worksheet = workbook[self.sheet_name] if self.sheet_name else workbook.active

        records = list(worksheet.iter_rows(values_only=True))
        if not records:
            return []

        headers = [str(value).strip() if value is not None else "" for value in records[0]]
        normalized_headers = [_normalize_header(h) for h in headers]

        rows: List[Dict[str, str]] = []
        for values in records[1:]:
            if not any(v is not None and str(v).strip() for v in values):
                continue
            row: Dict[str, str] = {}
            for idx, header in enumerate(normalized_headers):
                if not header:
                    continue
                raw_value = values[idx] if idx < len(values) else ""
                row[header] = str(raw_value).strip() if raw_value is not None else ""
            if row:
                rows.append(row)
        return rows

    def _read_csv_rows(self) -> List[Dict[str, str]]:
        with open(self.table_path, "r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            rows: List[Dict[str, str]] = []
            for raw_row in reader:
                row: Dict[str, str] = {}
                for key, value in raw_row.items():
                    if key is None:
                        continue
                    row[_normalize_header(key)] = str(value or "").strip()
                if any(v for v in row.values()):
                    rows.append(row)
        return rows

    def _build_matrix(self, rows: List[Dict[str, str]]) -> PromptMatrix:
        criteria_order: List[str] = []
        criteria_seen = set()
        pair_prompts: Dict[str, Dict[str, List[str]]] = {}

        if not rows:
            return PromptMatrix(criteria=[], pair_prompts={})

        headers = set(rows[0].keys())
        criterion_header = self._find_first(headers, _CRITERION_HEADER_ALIASES)
        technique_header = self._find_first(headers, _TECHNIQUE_HEADER_ALIASES)
        prompt_header = self._find_first(headers, _PROMPT_HEADER_ALIASES)

        # Long format: criterion + technique + prompt columns.
        if criterion_header and technique_header and prompt_header:
            for row in rows:
                criterion = _normalize_text(row.get(criterion_header, ""))
                technique = _normalize_technique(row.get(technique_header, ""))
                steps = _split_prompt_steps(row.get(prompt_header, ""))

                if not criterion or not technique or not steps:
                    continue

                self._add_pair_entry(
                    pair_prompts=pair_prompts,
                    criteria_order=criteria_order,
                    criteria_seen=criteria_seen,
                    criterion=criterion,
                    technique=technique,
                    steps=steps,
                )
            return PromptMatrix(criteria=criteria_order, pair_prompts=pair_prompts)

        # Wide format: criterion + one technique column per strategy.
        if not criterion_header:
            raise ValueError(
                "Prompt table must include a criterion/parameter column."
            )

        technique_columns: Dict[str, str] = {}
        for header in headers:
            normalized = _normalize_technique(header)
            if normalized and normalized not in technique_columns:
                technique_columns[normalized] = header

        if not technique_columns:
            raise ValueError(
                "Prompt table has no recognized technique columns. "
                f"Expected one of: {', '.join(SUPPORTED_TECHNIQUES)}"
            )

        for row in rows:
            criterion = _normalize_text(row.get(criterion_header, ""))
            if not criterion:
                continue

            for technique, header in technique_columns.items():
                steps = _split_prompt_steps(row.get(header, ""))
                if not steps:
                    continue
                self._add_pair_entry(
                    pair_prompts=pair_prompts,
                    criteria_order=criteria_order,
                    criteria_seen=criteria_seen,
                    criterion=criterion,
                    technique=technique,
                    steps=steps,
                )

        return PromptMatrix(criteria=criteria_order, pair_prompts=pair_prompts)

    def _find_first(self, headers: Iterable[str], candidates: set[str]) -> str:
        for header in headers:
            if header in candidates:
                return header
        return ""

    def _add_pair_entry(
        self,
        *,
        pair_prompts: Dict[str, Dict[str, List[str]]],
        criteria_order: List[str],
        criteria_seen: set[str],
        criterion: str,
        technique: str,
        steps: List[str],
    ) -> None:
        criterion_key = _canonical_criterion(criterion)
        if criterion_key not in criteria_seen:
            criteria_seen.add(criterion_key)
            criteria_order.append(criterion)

        if criterion_key not in pair_prompts:
            pair_prompts[criterion_key] = {}
        pair_prompts[criterion_key][technique] = steps
